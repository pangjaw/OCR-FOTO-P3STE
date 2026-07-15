"""merge_pdf_foto.py — Gabung foto hasil edit (format 2026) ke PDF 2025.

Flat structure: photos_dir/Tim_N/{btp}/{category}/{pdf_stem}/{0,50,100}.jpg
Output: output_dir/Tim_N/{btp}/{category}/{pdf_name}

No funcloc1 logic. Category detected from filename.
"""
import json
import os
import re
from pathlib import Path
from collections import defaultdict, Counter
from dataclasses import dataclass
import pdfplumber
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import argparse

from export_pdf_foto import (
    extract_station_from_description, load_sap_mapping, SAP_MAPPING_PATH,
    detect_category_from_filename, extract_station_from_filename, STATION_TO_BTP,
    extract_identifier, extract_funcloc_from_text, extract_all_funclocs,
)

DEFAULT_INPUT_DIR = "./02_pdf_target"
DEFAULT_PHOTOS_DIR = "./04_photos_edited"
DEFAULT_OUTPUT_DIR = "./05_pdf_merged"
CHECKLIST_CONFIG_PATH = "./checklist_types.json"


@dataclass
class AssetRow:
    page_number: int
    code: str
    title: str
    asset_type: str
    detail: str
    top: float
    station: str = "UNKNOWN"


# ── Helpers ─────────────────────────────────────────────────────

def log(msg):
    print(msg, flush=True)


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def sanitize_segment(text: str) -> str:
    text = normalize_spaces(text)
    text = "".join("_" if ord(char) < 32 else char for char in text)
    text = re.sub(r'[<>:\\\"/\\|?*]', "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text or "UNKNOWN"


def load_checklist_config(path: str = CHECKLIST_CONFIG_PATH) -> dict:
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        log(f"[WARNING] Checklist config not found at {path}, using defaults")
        return {"search_keyword": "PERAWATAN", "types": {}}


# ── Asset parsing ───────────────────────────────────────────────

def detect_asset_type(code: str, title: str) -> str:
    upper = f"{code} {title}".upper()
    if code.startswith("AXL") or "AXLE COUNTER" in upper:
        return "AXC"
    if code.startswith("WSL") or "WESEL" in upper:
        return "WESEL"
    if code.startswith("SIN") or "SINYAL" in upper:
        return "SINYAL"
    if code.startswith("CDA") or "CATU DAYA" in upper:
        return "CATU_DAYA"
    if code.startswith("JPL") or "PINTU PERLINTASAN" in upper:
        return "PINTU_PERLINTASAN"
    if code.startswith("TLK") or code.startswith("TWR") or "TELEKOMUNIKASI" in upper or "RADIO" in upper or "SERAT OPTIK" in upper or "OTB" in upper:
        return "TELEKOMUNIKASI"
    if code.startswith("CTC") or "CTC" in upper or "CTS" in upper or "DALWAS" in upper:
        return "CTS"
    if code.startswith("INB") or code.startswith("TRA"):
        return "UNKNOWN"
    if "BANGUNAN" in upper or "DATA LOGGER" in upper:
        return "PDSE"
    if "MULTIPLEX" in upper:
        return "PTLS"
    return "UNKNOWN"


def extract_detail(title: str, asset_type: str) -> str:
    original = normalize_spaces(title)

    def after(marker: str) -> str | None:
        match = re.search(re.escape(marker), original, flags=re.IGNORECASE)
        if not match:
            return None
        return normalize_spaces(original[match.end():]).lstrip(": -")

    detail = None
    if asset_type == "AXC":
        detail = after("COUNTER")
    elif asset_type == "WESEL":
        detail = after("ELEKTRIK")
    elif asset_type == "SINYAL":
        detail = after("ELEKTRIK") or after("SINYAL MUKA") or after("SINYAL")
    elif asset_type == "CATU_DAYA":
        detail = after("CATU DAYA") or after("GENSET") or after("UPS") or after("BATTERE") or after("BATT")
    elif asset_type == "PINTU_PERLINTASAN":
        detail = after("PINTU PERLINTASAN") or after("JPL") or after("JPLE") or after("GENTANIK")
    elif asset_type == "TELEKOMUNIKASI":
        detail = after("TELEKOMUNIKASI") or after("RADIO") or after("WAYSTATION") or after("SERAT OPTIK")
    elif asset_type == "CTS":
        detail = after("CTC") or after("PERALATAN") or original
    elif asset_type == "PERSINYALAN_ELEKTRIK":
        detail = after("PERSINYALAN ELEKTRIK") or after("DALAM PERSINYALAN") or after("OTB") or after("BANGUNAN")
    if not detail:
        detail = original
    sanitized = sanitize_segment(detail)
    if "ZP 41B" in sanitized.upper() or "ZP41B" in sanitized.upper():
        sanitized = re.sub(r'ZP\s*41B', 'ZP 41', sanitized, flags=re.IGNORECASE)
    return sanitized


def is_valid_asset_title(title: str) -> bool:
    words = title.split()
    code_pattern = re.compile(r"^[A-Z]{2,4}\d{4,}$")
    if all(code_pattern.match(w) for w in words):
        return False
    upper = title.upper()
    valid_keywords = [
        "AXLE", "COUNTER", "WESEL", "SINYAL", "CATU DAYA", "PINTU PERLINTASAN",
        "TELEKOMUNIKASI", "PERSINYALAN ELEKTRIK", "JPL", "GENTANIK", "RADIO",
        "SERAT OPTIK", "OTB", "BANGUNAN", "GENSET", "UPS", "BATTERE", "PANEL",
        "RECTIFIER", "MESIN", "MOTOR", "TOWER", "ANTENA", "INTERLOCKING", "INPUT",
    ]
    return any(kw in upper for kw in valid_keywords)


def extract_asset_rows(page: pdfplumber.page.Page, sap_mapping: dict = None) -> list[AssetRow]:
    lines = defaultdict(list)
    for word in page.extract_words(use_text_flow=True):
        lines[round(float(word["top"]), 1)].append(word)

    rows = []
    code_pattern = re.compile(r"^[A-Z]{2,4}\d{4,}$")

    for top in sorted(lines):
        words = sorted(lines[top], key=lambda w: float(w["x0"]))
        code_index = None
        code = None
        for idx, word in enumerate(words):
            if code_pattern.match(word["text"]):
                code_index = idx
                code = word["text"]
                break
        if code_index is None or not code:
            continue

        title = normalize_spaces(" ".join(word["text"] for word in words[code_index + 1:])).lstrip(": -")
        if not title or not is_valid_asset_title(title):
            continue

        asset_type = detect_asset_type(code, title)
        detail = extract_detail(title, asset_type)
        station = extract_station_from_description(title, sap_mapping or {}, code)
        rows.append(AssetRow(
            page_number=page.page_number, code=code, title=title,
            asset_type=asset_type, detail=detail, top=top, station=station,
        ))
    return rows


# ── PDF helpers ─────────────────────────────────────────────────

def extract_location_from_filename(filename: str) -> str:
    name = filename.rsplit('.', 1)[0]
    parts = name.split('_')
    if len(parts) >= 3:
        loc = parts[2].strip()
        loc = re.sub(r'\s*\(\d+\)\s*$', '', loc)
        return loc.upper()
    return "BOGOR"


def extract_date_from_page1(page: pdfplumber.page.Page) -> str:
    text = page.extract_text() or ""
    m = re.search(r"Tanggal\s*:\s*(\d{4}-\d{2}-\d{2})", text, re.IGNORECASE)
    months_id = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember",
    }
    if m:
        y, mo, d = map(int, m.group(1).split('-'))
        return f"{d:02d} {months_id[mo]} {y}"
    return "06 Januari 2025"


def extract_checklist_title(page: pdfplumber.page.Page, filename: str, config: dict | None = None) -> str:
    if config is None:
        config = load_checklist_config()
    search_keyword = config.get("search_keyword", "PERAWATAN")
    known_types = config.get("types", {})

    text = page.extract_text() or ""
    for line in text.split('\n'):
        line_clean = normalize_spaces(line)
        if search_keyword in line_clean.upper():
            if line_clean.upper().startswith("STE"):
                line_clean = line_clean[3:].strip()
            extracted = line_clean.upper()
            for known_type in known_types:
                if known_type in extracted or extracted in known_type:
                    return known_type
            return extracted
    # Fallback from filename
    name = filename.rsplit('.', 1)[0]
    parts = name.split('_')
    if len(parts) >= 2:
        ft = parts[1].strip().upper()
        for known_type in known_types:
            if known_type in ft or ft in known_type:
                return known_type
        return ft
    return "PERAWATAN AXLE COUNTER SIEMENS 1 BULANAN"


def get_text_width(text: str, fontname: str, fontsize: float) -> float:
    return fitz.get_text_length(text, fontname=fontname, fontsize=fontsize)


def draw_centered_text(page, text: str, y_baseline: float, fontname: str, fontsize: float):
    w = get_text_width(text, fontname, fontsize)
    x = (page.rect.width - w) / 2
    page.insert_text((x, y_baseline), text, fontname=fontname, fontsize=fontsize, color=(0, 0, 0))


def draw_centered_label(page, text: str, img_x0: float, img_x1: float, y_baseline: float, fontname: str, fontsize: float):
    w = get_text_width(text, fontname, fontsize)
    center_img = (img_x0 + img_x1) / 2
    x = center_img - w / 2
    page.insert_text((x, y_baseline), text, fontname=fontname, fontsize=fontsize, color=(0, 0, 0))


def draw_header(page, location: str, date_str: str, checklist_title: str):
    draw_centered_text(page, "FOTO DOKUMENTASI", 38.9, "hebo", 7.2)
    draw_centered_text(page, f"{checklist_title} {location}", 53.3, "hebo", 7.2)
    draw_centered_text(page, date_str, 67.7, "hebo", 7.2)


def determine_btp_from_identifier(identifier: str) -> str:
    """Determine BTP from any identifier format (handles all categories)."""
    if not identifier:
        return "BTP JAK"

    if identifier.startswith("JPL "):
        codes = re.findall(r'\b(BOO|CLT|BJD|BOP|BTT|CGB|CS|COS|MSG|CCR)\b', identifier.upper())
        if codes:
            station = codes[0].upper()
            if station == "CS": station = "COS"
            return STATION_TO_BTP.get(station, "BTP JAK")
        return "BTP JAK"
    elif identifier.startswith("ER ") or identifier.startswith("RUANG ") or identifier.startswith("RADIO_"):
        parts = identifier.split()
        code = None
        for kw in ["BOO", "BTT", "CLT", "BOP", "CGB", "COS", "MSG"]:
            if kw in parts:
                code = kw
                break
        if not code: code = "BOO"
        return STATION_TO_BTP.get(code, "BTP JAK")
    else:
        # Generic: extract last uppercase word as station code
        # Works for: "W31D BOO", "J10 BOO", "ZP 201B MSG", "BOO", etc.
        codes = re.findall(r'\b(BOO|CLT|BJD|BOP|BTT|CGB|CS|COS|MSG|CCR)\b', identifier.upper())
        if codes:
            station = codes[-1].upper()
            if station == "CS": station = "COS"
            return STATION_TO_BTP.get(station, "BTP JAK")
        return "BTP JAK"


# ── Core Merge ──────────────────────────────────────────────────

def _iter_tim_dirs(photos_dir: Path, pdf_tim: int | None):
    """Yield Tim_N directories to search, prioritizing schedule tim."""
    if pdf_tim:
        d = photos_dir / f"Tim_{pdf_tim}"
        if d.is_dir():
            yield d
    if photos_dir.exists():
        for d in sorted(photos_dir.iterdir()):
            if d.is_dir() and d.name.startswith("Tim_") and (not pdf_tim or d.name != f"Tim_{pdf_tim}"):
                yield d


def process_pdf(pdf_path: Path, photos_dir: Path, output_dir: Path,
                 input_root: Path | None = None, schedule_lookup: dict | None = None,
                 sap_mapping: dict | None = None) -> str:
    """Merge edited photos into PDF.
    
    Flat structure. Photo lookup via Funcloc identifier from page 1.
    Photo path: photos_dir/Tim_N/{btp}/{category}/{identifier}/{0,50,100}.jpg
    Output: output_dir/Tim_N/{btp}/{category}/{pdf_name}
    """
    if sap_mapping is None:
        sap_mapping = {}
    
    # ── Detect category from filename ──
    category = detect_category_from_filename(pdf_path.name)

    # ── Load checklist config ──
    config = load_checklist_config()
    
    # ── Read page 1: extract ALL funclocs → identifiers ──
    with pdfplumber.open(str(pdf_path)) as plumber_pdf:
        if len(plumber_pdf.pages) == 0:
            return "failed: PDF has 0 pages"
        page1 = plumber_pdf.pages[0]
        page1_text = page1.extract_text() or ""
        date_str = extract_date_from_page1(page1)
        checklist_title = extract_checklist_title(page1, pdf_path.name, config)
        
        # Scan ALL funclocs on page 1
        all_funclocs = extract_all_funclocs(page1_text)
        print(f"  [MERGE] Funclocs on page 1: {len(all_funclocs)}")
        
        # Build list of (identifier, btp) pairs
        # Also auto-detect category from funcloc prefix if filename detection failed
        asset_entries = []
        seen = set()
        for funcloc_line in all_funclocs:
            identifier = extract_identifier(funcloc_line, category)
            if identifier and identifier not in seen:
                seen.add(identifier)
                btp = determine_btp_from_identifier(identifier)
                # Auto-detect category from funcloc prefix for photo lookup
                photo_category = category
                if category == "UNKNOWN":
                    txt = funcloc_line.strip().upper()
                    if txt.startswith("WSL"): photo_category = "WESEL"
                    elif txt.startswith("SIN"): photo_category = "SINYAL"
                    elif txt.startswith("AXL"): photo_category = "AXC"
                asset_entries.append((identifier, btp, photo_category))
                print(f"  [MERGE] '{funcloc_line.strip()[:80]}' -> '{identifier}' (BTP: {btp}, cat: {photo_category})")
        
        if not asset_entries:
            # Fallback: use pdf stem
            identifier = sanitize_segment(pdf_path.stem)
            btp = "BTP JAK"
            asset_entries.append((identifier, btp, category))
            print(f"  [MERGE] Fallback identifier from stem: '{identifier}'")
    
    # ── Determine Tim ──
    pdf_tim = None
    if schedule_lookup:
        pdf_tim = schedule_lookup.get(pdf_path.name)
    
    location = extract_location_from_filename(pdf_path.name)
    
    # ── Find photos for all identifiers ──
    photo_paths = {}  # identifier -> (0.jpg, 50.jpg, 100.jpg)
    missing = []
    
    for identifier, btp, photo_category in asset_entries:
        found = False
        
        # WESEL: use glob to find base + date-suffixed folders
        if photo_category == "WESEL":
            for tim_dir in _iter_tim_dirs(photos_dir, pdf_tim):
                base_dir = tim_dir / btp / "WESEL"
                if not base_dir.is_dir():
                    continue
                # Glob: matches "W11 CLT", "W11 CLT_05-01", "W11 CLT_17-01"
                matching = sorted(d for d in base_dir.glob(f"{identifier}*") if d.is_dir())
                for d in matching:
                    f0 = d / "0.jpg"
                    f50 = d / "50.jpg"
                    f100 = d / "100.jpg"
                    if f0.is_file() and f50.is_file() and f100.is_file():
                        photo_paths[d.name] = (f0, f50, f100)
                        found = True
            # Also try direct (no Tim_N prefix)
            direct = photos_dir / btp / "WESEL"
            if direct.is_dir():
                for d in sorted(direct.glob(f"{identifier}*")):
                    if d.is_dir():
                        f0, f50, f100 = d/"0.jpg", d/"50.jpg", d/"100.jpg"
                        if all(f.is_file() for f in [f0, f50, f100]):
                            photo_paths[d.name] = (f0, f50, f100)
                            found = True
            if not found:
                missing.append(f"{identifier} ({btp}/{photo_category})")
            continue

        # Non-WESEL: original logic
        search_dirs = []
        if pdf_tim:
            search_dirs.append(photos_dir / f"Tim_{pdf_tim}" / btp / photo_category / identifier)
        
        # Scan all Tim_N directories
        if photos_dir.exists():
            for tim_dir in sorted(photos_dir.iterdir()):
                if tim_dir.is_dir() and tim_dir.name.startswith("Tim_"):
                    search_dirs.append(tim_dir / btp / photo_category / identifier)
        
        # Also try direct (no Tim_N prefix)
        search_dirs.append(photos_dir / btp / photo_category / identifier)
        
        for search_dir in search_dirs:
            f0 = search_dir / "0.jpg"
            f50 = search_dir / "50.jpg"
            f100 = search_dir / "100.jpg"
            if f0.is_file() and f50.is_file() and f100.is_file():
                photo_paths[identifier] = (f0, f50, f100)
                found = True
                break
        
        if not found:
            # Second pass: search ALL BTP folders under each Tim_N
            if photos_dir.exists():
                for tim_dir in sorted(photos_dir.iterdir()):
                    if tim_dir.is_dir() and tim_dir.name.startswith("Tim_"):
                        for btp_dir in sorted(tim_dir.iterdir()):
                            if btp_dir.is_dir():
                                search_dir = btp_dir / photo_category / identifier
                                f0 = search_dir / "0.jpg"
                                f50 = search_dir / "50.jpg"
                                f100 = search_dir / "100.jpg"
                                if f0.is_file() and f50.is_file() and f100.is_file():
                                    photo_paths[identifier] = (f0, f50, f100)
                                    found = True
                                    break
                        if found:
                            break
        
        if not found:
            missing.append(f"{identifier} ({btp}/{photo_category})")
    
    if not photo_paths:
        # No photos found — copy PDF to output without photo pages, removing last page
        doc = fitz.open(str(pdf_path))
        if len(doc) > 0:
            doc.delete_page(-1)
        first_btp = asset_entries[0][1] if asset_entries else "BTP JAK"
        first_category = asset_entries[0][2] if asset_entries else category
        out_pdf_path = output_dir / f"Tim_{pdf_tim or 1}" / first_btp / first_category / pdf_path.name
        ensure_dir(out_pdf_path.parent)
        doc.save(str(out_pdf_path))
        doc.close()
        print(f"  [MERGE] Saved (no photos): {out_pdf_path}")
        return "ok"
    
    if missing:
        print(f"  [MERGE] Missing photos: {', '.join(missing)}")
    
    # ── Build merged PDF ──
    doc = fitz.open(str(pdf_path))
    
    # Remove last page (old photo collage)
    if len(doc) > 0:
        doc.delete_page(-1)
    
    # Create new photo pages
    assets_per_page = 4
    entry_list = list(photo_paths.items())
    page_count_after = 0
    
    for i, (identifier, (f0_path, f50_path, f100_path)) in enumerate(entry_list):
        page_idx = i // assets_per_page
        asset_idx_on_page = i % assets_per_page
        
        if asset_idx_on_page == 0:
            page = doc.new_page(width=595, height=842)
            draw_header(page, location, date_str, checklist_title)
            page_count_after += 1
        
        page = doc[-1]
        
        # Asset title
        y_title_base = 82.1 + asset_idx_on_page * 183
        page.insert_text((31.5, y_title_base), identifier, fontname="helv", fontsize=7.2, color=(0, 0, 0))
        
        # Photos
        y_img_top = 89.1 + asset_idx_on_page * 183
        y_img_bottom = y_img_top + 148.8
        
        rect_col0 = fitz.Rect(31.5, y_img_top, 180.3, y_img_bottom)
        page.insert_image(rect_col0, filename=str(f0_path))
        
        rect_col1 = fitz.Rect(210.4, y_img_top, 359.2, y_img_bottom)
        page.insert_image(rect_col1, filename=str(f50_path))
        
        rect_col2 = fitz.Rect(389.8, y_img_top, 538.6, y_img_bottom)
        page.insert_image(rect_col2, filename=str(f100_path))
        
        # Labels
        y_label_base = 251.3 + asset_idx_on_page * 183
        draw_centered_label(page, "Foto 0%", 31.5, 180.3, y_label_base, "helv", 7.2)
        draw_centered_label(page, "Foto 50%", 210.4, 359.2, y_label_base, "helv", 7.2)
        draw_centered_label(page, "Foto 100%", 389.8, 538.6, y_label_base, "helv", 7.2)
    
    # ── Save output ──
    first_btp = asset_entries[0][1] if asset_entries else "BTP JAK"
    first_category = asset_entries[0][2] if asset_entries else category
    
    if pdf_tim:
        out_pdf_path = output_dir / f"Tim_{pdf_tim}" / first_btp / first_category / pdf_path.name
    else:
        out_pdf_path = output_dir / f"Tim_1" / first_btp / first_category / pdf_path.name
    
    ensure_dir(out_pdf_path.parent)
    
    if os.environ.get("OVERWRITE", "1") == "0" and out_pdf_path.exists():
        doc.close()
        return f"skipped: {out_pdf_path.name} sudah ada (overwrite=off)"
    
    doc.save(str(out_pdf_path))
    doc.close()
    print(f"  [MERGE] Saved: {out_pdf_path} ({len(entry_list)} assets)")
    
    return "ok"


# ── Main ────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Gabungkan foto hasil edit ke PDF (flat structure).")
    p.add_argument("--input", default=DEFAULT_INPUT_DIR)
    p.add_argument("--photos", default=DEFAULT_PHOTOS_DIR)
    p.add_argument("--output", default=DEFAULT_OUTPUT_DIR)
    p.add_argument("--schedule", default=None, help="schedule.json for Tim lookup")
    return p.parse_args()


def main():
    sap_mapping = load_sap_mapping(SAP_MAPPING_PATH)
    args = parse_args()
    input_dir = Path(args.input).resolve()
    photos_dir = Path(args.photos).resolve()
    output_dir = Path(args.output).resolve()

    if not input_dir.is_dir():
        print(f"Folder input tidak ditemukan: {input_dir}")
        return 1
    if not photos_dir.is_dir():
        print(f"Folder foto hasil edit tidak ditemukan: {photos_dir}")
        return 1

    ensure_dir(output_dir)

    pdf_files = sorted([p for p in input_dir.rglob("*")
                        if p.is_file() and p.suffix.lower() == ".pdf"])
    if not pdf_files:
        print(f"Tidak ada berkas PDF di: {input_dir}")
        return 0

    # Load schedule
    schedule_lookup = None
    if args.schedule:
        sched_path = Path(args.schedule)
        if sched_path.exists():
            with open(sched_path, encoding="utf-8") as f:
                sched_data = json.load(f)
            schedule_lookup = {e["file"]: e["tim"] for e in sched_data.get("schedules", [])}
            print(f"[SCHEDULE] Loaded {len(schedule_lookup)} file->Tim mappings.")

    log(f"Mulai pemrosesan {len(pdf_files)} berkas PDF...")
    log(f"Input:  {input_dir}")
    log(f"Photos: {photos_dir}")
    log(f"Output: {output_dir}\n")

    success = 0
    skipped = 0
    failed = 0
    failed_files = []
    skipped_files = []

    for pdf_path in pdf_files:
        status = process_pdf(pdf_path, photos_dir, output_dir, input_dir, schedule_lookup, sap_mapping)
        if status == "ok":
            success += 1
            log(f"[OK] {pdf_path.name}")
        elif status.startswith("skipped"):
            skipped += 1
            skipped_files.append({
                "file": pdf_path.name, "reason": status,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            log(f"[SKIP] {pdf_path.name} - {status}")
        else:
            failed += 1
            failed_files.append({
                "file": pdf_path.name, "reason": status,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            log(f"[FAIL] {pdf_path.name} - {status}")

    # Export Excel logs
    logs_dir = Path("logs")
    ensure_dir(logs_dir)

    if failed_files:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Failed Files"
        for col, h in enumerate(["File PDF", "Alasan", "Timestamp"], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
        for row_idx, item in enumerate(failed_files, 2):
            ws.cell(row=row_idx, column=1, value=item["file"])
            ws.cell(row=row_idx, column=2, value=item["reason"])
            ws.cell(row=row_idx, column=3, value=item["timestamp"])
        for col in ws.columns:
            ml = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = ml + 2
        wb.save(logs_dir / "merge_failed.xlsx")
        log(f"[EXPORT] Failed files -> logs/merge_failed.xlsx ({len(failed_files)})")

    if skipped_files:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Skipped Files"
        for col, h in enumerate(["File PDF", "Alasan", "Timestamp"], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        for row_idx, item in enumerate(skipped_files, 2):
            ws.cell(row=row_idx, column=1, value=item["file"])
            ws.cell(row=row_idx, column=2, value=item["reason"])
            ws.cell(row=row_idx, column=3, value=item["timestamp"])
        for col in ws.columns:
            ml = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = ml + 2
        wb.save(logs_dir / "merge_skipped.xlsx")
        log(f"[EXPORT] Skipped files -> logs/merge_skipped.xlsx ({len(skipped_files)})")

    summary = {
        "step": "merge",
        "success": success,
        "failed": failed,
        "skipped": skipped,
        "failed_file": "logs/merge_failed.xlsx" if failed_files else None,
        "skipped_file": "logs/merge_skipped.xlsx" if skipped_files else None,
    }
    print(f"__SUMMARY__:{json.dumps(summary)}", flush=True)
    log(f"\nSelesai. Sukses: {success}, Dilewati: {skipped}, Gagal: {failed}.")
    return 0 if success > 0 or skipped > 0 else 1


if __name__ == "__main__":
    import sys
    sys.exit(main())
