#!/usr/bin/env python3
"""
edit_timemark_ide1.py — Versi baru: HSV Orange Isolation + Fixed-offset Stage 1c
- Hapus Stage 1, 1b, 2, 3, 4, consensus pre-scan
- Hanya: HSV isolate -> find_red_guide -> fixed offset dari guide top-right
- Fallback: get_text_box() jika tidak ada guide
"""

import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
from PIL import Image, ImageDraw, ImageFont, ImageOps, ImageFilter

import pytesseract

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

# Fixed offset parameters (dari test script)
X_OFFSET_FROM_GUIDE = 6      # px right of guide right edge
Y_CENTER_OFFSET = 0          # textbox center = guide top (gy1)
BOX_HEIGHT_RATIO = 0.053     # 5.3% of image height
BOX_WIDTH_RATIO = 0.42       # 42% of image width

# Default paths
DEFAULT_INPUT = Path("03_photos_export")
DEFAULT_OUTPUT = Path("04_photos_edited")
DEFAULT_SCHEDULE = Path("schedule.json")

# ── Fallback durations (minutes) when schedule.json is not available ──
CATEGORY_DURATIONS = {
    "SINYAL": 30,
    "PERSINYALAN_ELEKTRIK": 420,
    "TELEKOMUNIKASI": 60,
    "SERAT OPTIK": 60,
    "PDSE": 420,
}
DEFAULT_DURATION = 45  # AXC, WESEL, CATU_DAYA, PINTU_PERLINTASAN, CTS, PTDS, PTLS, PTPP, JPL

MONTHS_ABBR = ["", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
DAYS_ID = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]


def parse_date_text(text: str) -> datetime:
    """Parse Indonesian date text like 'Kamis, Jan 09 2025' or 'Kamis, Jan 09 2025 07:00'."""
    m = re.match(r'\w+, (\w+) (\d{2}) (\d{4})(?: (\d{2}):(\d{2}))?', text)
    if m:
        month = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"Mei":5,"Jun":6,
                  "Jul":7,"Agu":8,"Sep":9,"Okt":10,"Nov":11,"Des":12}.get(m.group(1), 1)
        return datetime(int(m.group(3)), month, int(m.group(2)),
                        int(m.group(4)) if m.group(4) else 7,
                        int(m.group(5)) if m.group(5) else 0)
    return datetime.now()


def format_timemark(dt: datetime) -> str:
    return f"{DAYS_ID[dt.weekday()]}, {MONTHS_ABBR[dt.month]} {dt.day:02d} {dt.year} {dt.hour:02d}:{dt.minute:02d}"

# Station → BTP mapping (must match export_pdf_foto.py)
STATION_TO_BTP = {
    "BOO": "BTP JAK", "CLT": "BTP JAK",
    "BJD": "BTP JAK",
    "BNR": "BOP-BTT",
    "BOP": "BTP BD", "BTT": "BTP BD", "CGB": "BTP BD",
    "COS": "BTP BD", "MSG": "BTP BD",
}


# ─── HSV Orange Isolation ───
def preprocess_for_guide(img: Image.Image) -> np.ndarray:
    """Grayscale all colors EXCEPT pure red guide (Hue 0-15° or 350-360°, S>60%, V 30-80%)."""
    arr = np.array(img).astype(np.float32) / 255.0
    r, g, b = arr[:,:,0], arr[:,:,1], arr[:,:,2]

    maxc = np.maximum(np.maximum(r, g), b)
    minc = np.minimum(np.minimum(r, g), b)
    val = maxc
    delta = maxc - minc
    # Suppress divide-by-zero warning; np.where handles maxc==0 case
    with np.errstate(invalid='ignore', divide='ignore'):
        sat = np.where(maxc > 0, delta / maxc, 0)

    hue = np.zeros_like(r)
    m = delta > 0
    rm = m & (maxc == r)
    hue[rm] = (60 * ((g[rm] - b[rm]) / delta[rm])) % 360
    gm = m & (maxc == g)
    hue[gm] = 60 * ((b[gm] - r[gm]) / delta[gm]) + 120
    bm = m & (maxc == b)
    hue[bm] = 60 * ((r[bm] - g[bm]) / delta[bm]) + 240

    # Pure red mask: Hue 0-10° or 350-360°, Sat>65%, Val 25-85%
    # Excludes orange (15-35°) that appears behind red guide
    is_red = (
        ((hue <= 10) | (hue >= 350))
        & (sat >= 0.65)
        & (val >= 0.25)
        & (val <= 0.85)
    )

    gray = 0.299*r + 0.587*g + 0.114*b

    out = np.stack([
        np.where(is_red, r, gray),
        np.where(is_red, g, gray),
        np.where(is_red, b, gray),
    ], axis=2)

    return (np.clip(out, 0, 1) * 255).astype(np.uint8)


# ─── Red Guide Detection ───
def contiguous_runs(values: np.ndarray):
    idx = np.where(values)[0]
    if len(idx) == 0:
        return []
    groups = []
    s = p = int(idx[0])
    for v in idx[1:]:
        v = int(v)
        if v > p + 1:
            groups.append((s, p))
            s = v
        p = v
    groups.append((s, p))
    return groups


def find_red_guide(arr: np.ndarray):
    """Deteksi garis merah/oren vertikal (Red Guide) di kiri bawah."""
    h, w = arr.shape[:2]
    r = arr[:,:,0].astype(int)
    g = arr[:,:,1].astype(int)
    b = arr[:,:,2].astype(int)

    red = (
        (r > 145)
        & (r > g * 1.5)
        & (r > b * 1.5)
        & (g < 195)
        & (b < 155)
    )
    red[:int(h * 0.54), :] = False
    red[:, int(w * 0.16):] = False

    counts = red.sum(axis=0)
    min_count = max(20, int(h * 0.04))
    if not np.any(counts > min_count):
        return None

    candidates = []
    for x1, x2 in contiguous_runs(counts > min_count):
        if x2 - x1 + 1 > int(w * 0.08):  # relaxed for HSV-isolated wider borders
            continue
        row_has = red[:, x1:x2+1].sum(axis=1) > 0
        for y1, y2 in contiguous_runs(row_has):
            guide_h = y2 - y1 + 1
            if guide_h < max(25, int(h * 0.055)):
                continue
            score = int(red[y1:y2+1, x1:x2+1].sum())
            candidates.append((guide_h, score, x1, y1, x2, y2))

    if not candidates:
        return None
    _, _, x1, y1, x2, y2 = max(candidates)
    return x1, y1, x2, y2


def place_textbox_fixed_offset(guide: tuple[int,int,int,int], w: int, h: int) -> tuple[int,int,int,int]:
    """Place textbox at fixed offset from guide top-right."""
    _, gy1, gx2, _ = guide
    
    box_h = int(h * BOX_HEIGHT_RATIO)
    box_w = int(w * BOX_WIDTH_RATIO)
    
    new_x1 = gx2 + X_OFFSET_FROM_GUIDE
    new_y1 = gy1 + Y_CENTER_OFFSET - box_h // 2
    new_x2 = min(w, new_x1 + box_w)
    new_y2 = new_y1 + box_h
    
    # Clamp
    new_y1 = max(0, new_y1)
    new_y2 = min(h, new_y2)
    new_x1 = max(0, min(new_x1, w - 1))
    new_x2 = min(w, new_x2)
    
    return new_x1, new_y1, new_x2, new_y2


def get_text_box(w: int, h: int) -> tuple[int, int, int, int]:
    """Fallback area when no guide detected (bottom-left watermark position)."""
    x1 = int(w * 0.047)
    y1 = int(h * 0.735)
    x2 = int(w * 0.49)
    y2 = int(h * 0.82)
    return x1, y1, x2, y2


# ─── Diffuse Fill (inpainting) ───
def diffuse_fill(arr: np.ndarray, mask: np.ndarray, steps: int = 60) -> np.ndarray:
    filled = arr.copy().astype(np.float32)
    mask3 = mask[:, :, None]
    for _ in range(steps):
        up = np.vstack([filled[:1], filled[:-1]])
        down = np.vstack([filled[1:], filled[-1:]])
        left = np.hstack([filled[:, :1], filled[:, :-1]])
        right = np.hstack([filled[:, 1:], filled[:, -1:]])
        avg = (up + down + left + right) * 0.25
        filled = np.where(mask3, avg, filled)
    return np.clip(filled, 0, 255).astype(np.uint8)


# ─── Font ───
def get_font(size: int):
    candidates = [
        Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "arial.ttf",
        Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "segoeui.ttf",
        Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "calibri.ttf",
        "DejaVuSans.ttf",
        "DejaVuSans-Bold.ttf",
    ]
    for fp in candidates:
        if isinstance(fp, str):
            try:
                return ImageFont.truetype(fp, size=size)
            except:
                continue
        if fp.exists():
            return ImageFont.truetype(str(fp), size=size)
    return ImageFont.load_default()


# ─── Draw Textbox ───
def draw_textbox(img: Image.Image, box: tuple[int,int,int,int], text: str) -> Image.Image:
    x1, y1, x2, y2 = box
    w, h = img.size
    box_h = y2 - y1
    font_size = max(9, min(160, int(w * 0.038)))
    font = get_font(font_size)

    dummy = ImageDraw.Draw(Image.new("L", (1,1)))
    fit_stroke = 1 if font_size >= 28 else 0
    bbox = dummy.textbbox((0,0), text, font=font, stroke_width=fit_stroke)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    pad_x = max(2, int(w * 0.008))
    pad_y = max(1, int(h * 0.004))
    tx = x1 + pad_x
    ty = y1 + ((box_h - th) // 2) - bbox[1]

    shape_pad_x = max(3, int(font_size * 0.28))
    shape_pad_y = max(1, int(font_size * 0.12))
    sx1 = max(0, tx + bbox[0] - shape_pad_x)
    sy1 = max(0, ty + bbox[1] - shape_pad_y)
    sx2 = min(w, tx + bbox[2] + shape_pad_x)
    sy2 = min(h, ty + bbox[3] + shape_pad_y)
    shape_box = (sx1, sy1, sx2, sy2)

    out = img.copy()
    draw = ImageDraw.Draw(out)

    radius = max(2, int(box_h * 0.25))
    overlay = Image.new("RGBA", out.size, (0,0,0,0))
    odraw = ImageDraw.Draw(overlay)
    odraw.rounded_rectangle(shape_box, radius=radius, fill=(0,0,0,140))
    out = Image.alpha_composite(out.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(out)

    so = max(1, int(font_size * 0.02))
    draw.text((tx+so, ty+so), text, font=font, fill=(35,35,35), stroke_width=fit_stroke, stroke_fill=(35,35,35))
    draw.text((tx, ty), text, font=font, fill=(248,248,248))

    return out


# (duplicate defs removed - see live versions below)


# ─── Main ───
def sanitize_segment(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', text).strip()


def asset_output_dir(root: Path, btp: str, category: str, identifier: str) -> Path:
    return root / sanitize_segment(btp) / category / sanitize_segment(identifier)


def _folder_key_from_path(rel: Path, input_dir: Path | None = None) -> tuple[str, str, str] | None:
    """Extract (btp, category, identifier) from relative path for flat structure.

    Structure: [BTP/]category/identifier/photo.jpg
    Returns: (btp, category, identifier) or None
    """
    parts = rel.parts
    btp_shift = 1 if parts and parts[0].startswith("BTP") else 0
    if len(parts) >= 3 + btp_shift:
        btp = parts[0] if btp_shift else "BTP JAK"
        return btp, parts[0 + btp_shift], parts[1 + btp_shift]
    return None






# ─── Core: Locate Date Box ───
def locate_date_box(arr_orig: np.ndarray, arr_hsv: np.ndarray, w: int, h: int,
                    y_override: int | None = None, consensus_gy1: int | None = None):
    """
    Single-stage detection:
    1. If y_override -> use it
    2. Original image guide -> fixed offset textbox
    3. Folder consensus gy1 -> fixed offset textbox
    4. Fallback -> get_text_box()
    Returns: (box, stage_name)
    """
    if y_override is not None:
        box_h = int(h * BOX_HEIGHT_RATIO)
        y1 = max(0, y_override - box_h // 2)
        y2 = min(h, y1 + box_h)
        x1 = int(w * 0.047)
        x2 = int(w * 0.49)
        return (x1, y1, x2, y2), "stage_0_override"

    # Use original image
    guide = find_red_guide(arr_orig)
    if guide is not None:
        new_box = place_textbox_fixed_offset(guide, w, h)
        return new_box, "stage_1c_guide_original"

    # Folder consensus gy1 fallback
    if consensus_gy1 is not None:
        synthetic_guide = (8, consensus_gy1, 9, consensus_gy1 + 84)
        new_box = place_textbox_fixed_offset(synthetic_guide, w, h)
        return new_box, "stage_1c_guide_consensus"

    # No guide at all -> fallback
    fallback = get_text_box(w, h)
    return fallback, "stage_fallback"


# ─── Schedule Support ───
def load_schedule(path: Path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def iso_to_timemark(iso_str: str) -> str:
    """2026-07-11T09:30:00 -> Rabu, Jul 11 2026 09:30"""
    dt = datetime.fromisoformat(iso_str)
    days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    months = ["", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
    return f"{days[dt.weekday()]}, {months[dt.month]} {dt.day:02d} {dt.year} {dt.hour:02d}:{dt.minute:02d}"


def build_schedule_lookup(schedule_data: dict):
    """Build {(btp, category, pdf_stem_or_identifier, photo_name): (timemark_text, tim_n)}

    Schedule v3 format: btp/category/[identifier]/photos
    Falls back to pdf_stem for v2.
    """
    lookup = {}
    for sched in schedule_data.get("schedules", []):
        tim_n = sched.get("tim", 1)
        btp = sched.get("btp", "BTP JAK")
        category = sched.get("category", "UNKNOWN")
        identifier = sched.get("identifier") or sched.get("pdf_stem", "")
        for pname, iso_ts in sched.get("photos", {}).items():
            lookup[(btp, category, identifier, pname)] = (iso_to_timemark(iso_ts), tim_n)
    return lookup


# ─── Process Single Image ───
def process_image(image_path: Path, output_path: Path, date_text: str,
                  y_override: int | None = None, schedule_lookup: dict | None = None,
                  asset_key: tuple | None = None, consensus_gy1: int | None = None):
    img_orig = Image.open(image_path).convert('RGB')
    orig_w, orig_h = img_orig.size
    arr_orig = np.array(img_orig)

    # HSV isolation (kept for compatibility, but locate_date_box uses original)
    arr_hsv = preprocess_for_guide(img_orig)

    # Determine timemark text
    timemark_text = date_text
    if schedule_lookup and asset_key:
        timemark_text = schedule_lookup.get(asset_key, (date_text, 1))[0]

    # Locate box - returns (box, stage)
    box, stage = locate_date_box(arr_orig, arr_hsv, orig_w, orig_h, y_override, consensus_gy1)
    x1, y1, x2, y2 = box

    # Erase: precision mask (rounded rect, 1px pad)
    pad = 1
    ex1 = max(0, x1 - pad)
    ey1 = max(0, y1 - pad)
    ex2 = min(orig_w, x2 + pad)
    ey2 = min(orig_h, y2 + pad)

    crop = arr_orig[ey1:ey2, ex1:ex2]
    box_h_px = y2 - y1
    radius = max(2, int(box_h_px * 0.25))

    mask_img = Image.new("L", (crop.shape[1], crop.shape[0]), 0)
    md = ImageDraw.Draw(mask_img)
    md.rounded_rectangle(
        (x1 - ex1, y1 - ey1, x2 - ex1, y2 - ey1),
        radius=radius, fill=255
    )
    crop_mask = np.array(mask_img) > 0

    arr_work = arr_orig.copy()
    arr_work[ey1:ey2, ex1:ex2] = diffuse_fill(crop, crop_mask, steps=60)

    img_erased = Image.fromarray(arr_work)

    # Draw new textbox
    img_out = draw_textbox(img_erased, box, timemark_text)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    img_out.save(output_path, quality=95, subsampling=0)

    return stage


def _collect_folder_consensus(input_dir: Path, all_jpgs: list[Path]) -> dict[tuple[str, str, str], int]:
    from collections import defaultdict
    folder_guides = defaultdict(list)

    for src in all_jpgs:
        rel = src.relative_to(input_dir)
        fkey = _folder_key_from_path(rel, input_dir)
        if not fkey:
            continue
        try:
            img = Image.open(src).convert('RGB')
            arr = np.array(img)
            guide = find_red_guide(arr)
            if guide:
                gx1, gy1, gx2, gy2 = guide
                folder_guides[fkey].append(gy1)
        except Exception:
            pass

    consensus = {}
    for fkey, gy1s in folder_guides.items():
        if len(gy1s) >= 2:
            gy1s.sort()
            median = gy1s[len(gy1s) // 2]
            if all(abs(g - median) <= 10 for g in gy1s):
                consensus[fkey] = median
    return consensus


def main():
    parser = argparse.ArgumentParser(description="Edit timemark on photos (Fixed-offset + Folder Consensus)")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Input photos folder")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output photos folder")
    parser.add_argument("--date", type=str, help="Global date text (e.g. 'Sabtu, Apr 29 2026 08:00')")
    parser.add_argument("--schedule", type=Path, help="schedule.json for per-photo timestamps")
    parser.add_argument("--y-override", type=int, help="Force Y position for textbox")
    parser.add_argument("--clear-output", action="store_true", help="Clear output folder first")
    args = parser.parse_args()

    input_dir = args.input
    output_dir = args.output

    if args.clear_output and output_dir.exists():
        shutil.rmtree(output_dir)

    schedule_lookup = None
    if args.schedule:
        schedule_data = load_schedule(args.schedule)
        schedule_lookup = build_schedule_lookup(schedule_data)

    date_text = args.date or datetime.now().strftime("%A, %b %d %Y %H:%M").replace("Monday", "Senin").replace("Tuesday", "Selasa").replace("Wednesday", "Rabu").replace("Thursday", "Kamis").replace("Friday", "Jumat").replace("Saturday", "Sabtu").replace("Sunday", "Minggu")

    all_jpgs = list(input_dir.rglob("*.jpg"))
    folder_consensus = _collect_folder_consensus(input_dir, all_jpgs)

    ok = 0
    failed = 0
    stage_counts = {"stage_0_override": 0, "stage_1c_guide_original": 0, "stage_1c_guide_consensus": 0, "stage_fallback": 0}
    stage_details = []
    failed_files = []
    tim_mapping = {}  # photo_rel_path -> tim_n

    for src in all_jpgs:
        rel = src.relative_to(input_dir)
        parts = rel.parts
        # Flat structure: [BTP/]category/identifier/photo.jpg
        btp_shift = 1 if parts and parts[0].startswith("BTP") else 0
        if len(parts) >= 3 + btp_shift:
            btp = parts[0] if btp_shift else "BTP JAK"
            category = parts[0 + btp_shift]
            identifier = parts[1 + btp_shift]
            photo_name = parts[2 + btp_shift]
            asset_key = (btp, category, identifier, photo_name)
        else:
            btp, category, identifier, photo_name, asset_key = "BTP JAK", "UNKNOWN", "", "", None

        tim_n = 1
        if schedule_lookup and asset_key:
            timemark_text, tim_n = schedule_lookup.get(asset_key, (date_text, 1))
            dst_dir = output_dir / f"Tim_{tim_n}" / asset_output_dir(Path(""), btp, category, identifier)
        else:
            # Fallback: compute time offset from photo name (0/50/100) * category duration
            pct = 0
            name_stem = Path(photo_name).stem
            if name_stem == "50":
                pct = 50
            elif name_stem == "100":
                pct = 100
            duration = CATEGORY_DURATIONS.get(category.upper(), DEFAULT_DURATION)
            offset_minutes = int(pct * duration / 100)
            base_dt = parse_date_text(date_text)
            photo_dt = base_dt + timedelta(minutes=offset_minutes)
            timemark_text = format_timemark(photo_dt)
            dst_dir = output_dir / f"Tim_{tim_n}" / asset_output_dir(Path(""), btp, category, identifier)
        tim_mapping[str(rel)] = tim_n
        dst = dst_dir / src.name

        fkey = _folder_key_from_path(rel, input_dir)
        consensus_gy1 = folder_consensus.get(fkey) if fkey else None

        try:
            stage = process_image(src, dst, timemark_text, args.y_override, schedule_lookup, asset_key, consensus_gy1)
            ok += 1
            if stage in stage_counts:
                stage_counts[stage] += 1
            stage_details.append({
                "file": str(rel),
                "stage": stage,
                "asset_type": category,
                "detail": identifier,
                "photo": photo_name,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            print(json.dumps({
                "type": "stage",
                "file": str(rel),
                "stage": stage,
                "asset_type": category,
                "detail": identifier,
                "photo": photo_name
            }), flush=True)
        except Exception as e:
            failed += 1
            failed_files.append({
                "file": str(rel),
                "asset_type": category,
                "detail": identifier,
                "photo": photo_name,
                "reason": str(e),
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })



    logs_dir = Path("logs")
    logs_dir.mkdir(parents=True, exist_ok=True)
    
    # Write tim_mapping.json if schedule was used
    if args.schedule and tim_mapping:
        tim_mapping_data = {
            "version": 1,
            "generated_at": datetime.now().isoformat(),
            "mapping": tim_mapping,
            "tims_used": sorted(set(tim_mapping.values()))
        }
        with open(logs_dir / "tim_mapping.json", "w", encoding="utf-8") as f:
            json.dump(tim_mapping_data, f, indent=2, ensure_ascii=False)

    if stage_details:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stage Details"
        headers = ["File", "Stage", "Asset Type", "Detail Aset", "Photo", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        for row_idx, item in enumerate(stage_details, 2):
            ws.cell(row=row_idx, column=1, value=item["file"])
            ws.cell(row=row_idx, column=2, value=item["stage"])
            ws.cell(row=row_idx, column=3, value=item["asset_type"])
            ws.cell(row=row_idx, column=4, value=item["detail"])
            ws.cell(row=row_idx, column=5, value=item["photo"])
            ws.cell(row=row_idx, column=6, value=item["timestamp"])
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        wb.save(logs_dir / "edit_stages.xlsx")

    if failed_files:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Failed Files"
        headers = ["File", "Asset Type", "Detail Aset", "Photo", "Alasan", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        for row_idx, item in enumerate(failed_files, 2):
            ws.cell(row=row_idx, column=1, value=item["file"])
            ws.cell(row=row_idx, column=2, value=item["asset_type"])
            ws.cell(row=row_idx, column=3, value=item["detail"])
            ws.cell(row=row_idx, column=4, value=item["photo"])
            ws.cell(row=row_idx, column=5, value=item["reason"])
            ws.cell(row=row_idx, column=6, value=item["timestamp"])
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        wb.save(logs_dir / "edit_failed.xlsx")



    summary = {
        "step": "edit",
        "success": ok,
        "failed": failed,
        "stage_counts": stage_counts,
        "failed_file": "logs/edit_failed.xlsx" if failed_files else None,
        "stages_file": "logs/edit_stages.xlsx" if stage_details else None
    }
    print(f"__SUMMARY__:{json.dumps(summary)}", flush=True)


if __name__ == "__main__":
    main()
